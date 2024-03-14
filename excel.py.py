from openpyxl import load_workbook


wb = load_workbook('New Microsoft Excel Worksheet.xlsx')
sheet = wb.active


keyword = '<id>'

new_value = 'id'

for row in sheet.iter_rows(values_only=True):
    for cell_value in row:
        if cell_value == keyword:
            
            cell = sheet.cell(row=row[0], column=row.index(keyword)+1)

            cell.value = new_value


wb.save('New Microsoft Excel Worksheet.xlsx')
